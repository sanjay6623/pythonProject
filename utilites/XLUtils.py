import openpyxl
from openpyxl.styles import PatternFill


def getRowCount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return (sheet.max_row)

def getColumnCount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return (sheet.max_column)

def readData(file,sheetName,rownum,columnno):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return sheet.cell(row=rownum,column=columnno).value

def writeData(file,sheetName,rownum,columnno,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    sheet.cell(row=rownum,column=columnno).value =data
    workbook.save(file)

def fillGreen(file,sheetName,rownum,columnno,):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    fillgreen=PatternFill(start_color="60b212",end_color="60b212",fill_type="solid")
    sheet.cell(rownum,columnno).fill=fillgreen
    workbook.save(file)

def fillRed(file,sheetName,rownum,columnno,):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    fillred=PatternFill(start_color="ff0000",end_color="ff0000",fill_type="solid")
    sheet.cell(rownum,columnno).fill=fillred
    workbook.save(file)


